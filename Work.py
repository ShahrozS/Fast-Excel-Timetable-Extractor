import openpyxl
import re
import string
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook('Timetable.xlsx')
n = 0
sheets = wb.sheetnames
wsMonday = wb[sheets[17]]

# rows = wsf.iter_rows(min_row = 3,max_row = 51 , min_col =2 ,max_col =10)




def check(val):
    a = "BSE-4A" #this can be change to whatever class OR it can be taken as an argument from user.
    if a in val:    
        return True;
    else:
        return False;   



a = "BSE-4A"
print("MONDAY")
print()


print("Time ","     ","    Class","             ","Sub")

for row in range(5,51):
    for col in range(2,10):
        char = get_column_letter(col)
        
        if wsMonday[char+str(row)].value != None and check(wsMonday[char+str(row)].value):
            print(wsMonday[char+'3'].value  , "      ",wsMonday["A"+str(row)].value.strip(),"     ", wsMonday[char+str(row)].value ,"       ", end= " ")
            print() 
    
print()
print("WEDNESDAY")
print()

wsWednesday = wb[sheets[19]]


print("Time ","     ","    Class","             ","Sub")

for row in range(5,51):
    for col in range(2,10):
        char = get_column_letter(col)
        
        if wsWednesday[char+str(row)].value != None and check(wsWednesday[char+str(row)].value):
            print(wsWednesday[char+'3'].value  , "      ",wsWednesday["A"+str(row)].value.strip(),"     ", wsWednesday[char+str(row)].value ,"       ", end= " ")
            print() 
    
    
   
print()
print("THURSDAY")
print()
wsThursday= wb[sheets[20]]

print("Time ","     ","    Class","             ","Sub")

for row in range(5,51):
    for col in range(2,10):
        char = get_column_letter(col)
        
        if wsThursday[char+str(row)].value != None and check(wsThursday[char+str(row)].value):
            print(wsThursday[char+'3'].value  , "      ",wsThursday["A"+str(row)].value.strip(),"     ", wsThursday[char+str(row)].value ,"       ", end= " ")
            print() 
    

print()
print("FRIDAY")
print()

wsFriday = wb[sheets[21]]

print("Time ","     ","    Class","             ","Sub")

for row in range(5,51):
    for col in range(2,11):
        char = get_column_letter(col)
        
        if wsFriday[char+str(row)].value != None and check(wsFriday[char+str(row)].value):
            print(wsFriday[char+'3'].value  , "      ",wsFriday["A"+str(row)].value.strip(),"     ", wsFriday[char+str(row)].value ,"       ", end= " ")
            print() 
    
  
    
# for b,c,d,e,f,g,h,i in rows:

#     if b.value!=None and check(b.value):
#         char = get_column_letter()
#         print(wsf[char+"1"])
#         print(b.value) 
#     elif c.value!=None and check(c.value):
#         print(c.value)
#     elif d.value!=None and check(d.value):
#         print(d.value)
#     elif e.value!=None and check(e.value):
#         print(e.value)
#     elif f.value!=None and check(f.value):
#         print(f.value)
#     elif g.value!=None and check(g.value):
#         print(g.value)
#     elif h.value!=None and check(h.value):
#         print(h.value)
#     elif i.value!=None and check(i.value):
#         print(i.value)
    
    
